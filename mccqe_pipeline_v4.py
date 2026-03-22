#!/usr/bin/env python3
"""
MCCQE Study App - Session 1 Pipeline (Final Edition)
======================================================
- Loads existing OCR cache (skips re-OCR)
- Detects diagram/figure-heavy pages
- Vision-processes diagram pages for clinical meaning
- Extracts and saves diagram page images to ./images/
- Outputs image_index.json for GAS web app
- Generates full Excel study data

Usage:
    python mccqe_pipeline_v4.py \
        --pdf1 "Toronto Notes 2025.pdf" \
        --pdf2 "MCCQE-Part-I-practice-questions-and-answers-2025.pdf" \
        --cache "Toronto_Notes_2025_ocr_cache.txt" \
        --key YOUR_ANTHROPIC_API_KEY

Outputs:
    MCCQE_Study_Data.xlsx
    images/               (diagram page images, PNG)
    image_index.json      (maps each image to specialty/topic/page)
"""

import argparse, base64, io, json, os, re, sys, time
import fitz
import pytesseract
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ── install deps ──────────────────────────────────────────────────────────────
def install(pkg):
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

for pkg, imp in [("anthropic","anthropic"),("pymupdf","fitz"),
                  ("pytesseract","pytesseract"),("pillow","PIL"),
                  ("openpyxl","openpyxl"),("tqdm","tqdm")]:
    try:
        __import__(imp)
    except ImportError:
        print(f"Installing {pkg}...")
        install(pkg)

import anthropic

TESS = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.exists(TESS):
    pytesseract.pytesseract.tesseract_cmd = TESS

# ── config ────────────────────────────────────────────────────────────────────
TEXT_MODEL    = "claude-haiku-4-5-20251001"
VISION_MODEL  = "claude-sonnet-4-6"
MAX_TOKENS    = 3000
CHUNK_CHARS   = 9000
RETRY_DELAY   = 15
MAX_RETRIES   = 3
OCR_DPI       = 180
VISION_DPI    = 150
IMAGE_DPI     = 200   # resolution for saved diagram images

# Page is diagram-heavy if OCR word count below this
DIAGRAM_THRESHOLD = 80

SPECIALTIES = [
    "Cardiology","Respirology","Gastroenterology","Nephrology",
    "Hematology","Oncology","Endocrinology","Neurology",
    "Rheumatology","Infectious Disease","Dermatology",
    "Ophthalmology","Otolaryngology","Psychiatry","Obstetrics",
    "Gynecology","Pediatrics","Surgery","Orthopedics",
    "Emergency Medicine","Anesthesia","Radiology",
    "Family Medicine","Geriatrics","Pharmacology"
]

HEADER_FILLS = {
    "Flashcards":"1F497D","Bullet_Notes":"2E75B6",
    "MCQ":"375623","Dont_Miss":"7030A0"
}

# ── helpers ───────────────────────────────────────────────────────────────────
def chunk_text(text, size=CHUNK_CHARS):
    return [text[i:i+size] for i in range(0, len(text), size)]

def detect_specialty(text):
    tl = text.lower()
    for s in SPECIALTIES:
        if s.lower() in tl:
            return s
    return "General Medicine"

def safe_json(text):
    text = re.sub(r"```(?:json)?", "", text).strip().rstrip("`").strip()
    s, e = text.find("["), text.rfind("]")
    if s == -1 or e == -1:
        return []
    try:
        return json.loads(text[s:e+1])
    except:
        return []

def img_to_b64(img: Image.Image) -> str:
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.standard_b64encode(buf.getvalue()).decode()

def call_claude(client, system, user, model=None):
    m = model or TEXT_MODEL
    for attempt in range(MAX_RETRIES):
        try:
            r = client.messages.create(
                model=m, max_tokens=MAX_TOKENS,
                system=system,
                messages=[{"role":"user","content":user}]
            )
            return r.content[0].text
        except anthropic.RateLimitError:
            wait = RETRY_DELAY * (attempt + 1)
            print(f"\n  Rate limit. Waiting {wait}s...")
            time.sleep(wait)
        except Exception as ex:
            print(f"\n  API error: {ex}")
            time.sleep(RETRY_DELAY)
    return "[]"

# ── Phase 1: scan PDF for diagram pages ──────────────────────────────────────
def scan_for_diagram_pages(pdf_path: str) -> list:
    """Return list of (page_index, word_count) for diagram-heavy pages."""
    print("  Scanning for diagram/figure pages...")
    doc = fitz.open(pdf_path)
    diagram_pages = []
    for i in tqdm(range(len(doc)), desc="  Scanning"):
        page = doc[i]
        mat = fitz.Matrix(OCR_DPI / 72, OCR_DPI / 72)
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        text = pytesseract.image_to_string(img, lang="eng", config="--psm 6")
        wc = len(text.split())
        if wc < DIAGRAM_THRESHOLD:
            diagram_pages.append((i, wc))
    doc.close()
    print(f"  Found {len(diagram_pages)} diagram pages out of {len(doc)} total")
    return diagram_pages

# ── Phase 2: extract and save diagram images ──────────────────────────────────
def extract_diagram_images(pdf_path: str, diagram_pages: list,
                            client, out_dir: str = "images") -> list:
    """
    For each diagram page:
    - Save high-res PNG to out_dir
    - Send to Claude Vision for clinical description
    - Return image_index list
    """
    os.makedirs(out_dir, exist_ok=True)
    doc = fitz.open(pdf_path)
    image_index = []
    vision_texts = {}

    print(f"\n  Extracting {len(diagram_pages)} diagram images...")
    for page_idx, wc in tqdm(diagram_pages, desc="  Diagrams"):
        page = doc[page_idx]

        # Save high-res image
        mat_save = fitz.Matrix(IMAGE_DPI / 72, IMAGE_DPI / 72)
        pix_save = page.get_pixmap(matrix=mat_save)
        img_filename = f"page_{page_idx + 1:04d}.png"
        img_path = os.path.join(out_dir, img_filename)
        pix_save.save(img_path)

        # Vision processing for clinical content
        mat_vis = fitz.Matrix(VISION_DPI / 72, VISION_DPI / 72)
        pix_vis = page.get_pixmap(matrix=mat_vis)
        vis_img = Image.open(io.BytesIO(pix_vis.tobytes("png")))
        b64 = img_to_b64(vis_img)

        try:
            r = client.messages.create(
                model=VISION_MODEL,
                max_tokens=1000,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": "image/png",
                                "data": b64
                            }
                        },
                        {
                            "type": "text",
                            "text": (
                                "This is a page from Toronto Notes 2025 (MCCQE study material). "
                                "1. Identify what type of content this is (anatomy diagram, clinical algorithm, "
                                "drug table, diagnostic criteria, flowchart, etc.). "
                                "2. Extract ALL clinical information exactly as shown. "
                                "3. For diagrams: describe structures, labels, and relationships. "
                                "4. For tables: preserve all rows and columns with pipe separators. "
                                "5. For algorithms/flowcharts: describe decision points and pathways. "
                                "6. State the medical specialty this belongs to. "
                                "Format response as: "
                                "SPECIALTY: ...\nTYPE: ...\nTOPIC: ...\nCONTENT:\n[extracted content]"
                            )
                        }
                    ]
                }]
            )
            vision_text = r.content[0].text

            # Parse specialty and topic from Vision response
            specialty = detect_specialty(vision_text)
            topic = "General"
            type_match = re.search(r"TYPE:\s*(.+)", vision_text)
            topic_match = re.search(r"TOPIC:\s*(.+)", vision_text)
            spec_match  = re.search(r"SPECIALTY:\s*(.+)", vision_text)
            if topic_match:
                topic = topic_match.group(1).strip()
            if spec_match:
                sp = spec_match.group(1).strip()
                if sp in SPECIALTIES:
                    specialty = sp

            content_match = re.search(r"CONTENT:\n(.+)", vision_text, re.DOTALL)
            content = content_match.group(1).strip() if content_match else vision_text

            image_index.append({
                "page": page_idx + 1,
                "filename": img_filename,
                "specialty": specialty,
                "topic": topic,
                "type": type_match.group(1).strip() if type_match else "diagram",
                "description": content[:500]  # first 500 chars as preview
            })
            vision_texts[page_idx] = content
            time.sleep(0.5)

        except Exception as ex:
            print(f"\n  Vision failed page {page_idx+1}: {ex}")
            image_index.append({
                "page": page_idx + 1,
                "filename": img_filename,
                "specialty": detect_specialty(""),
                "topic": "Unknown",
                "type": "diagram",
                "description": ""
            })

    doc.close()
    return image_index, vision_texts

# ── Phase 3: merge OCR cache + Vision text ────────────────────────────────────
def build_merged_text(cache_path: str, vision_texts: dict) -> str:
    """Load OCR cache and inject Vision content for diagram pages."""
    print(f"\n  Loading OCR cache: {cache_path}")
    with open(cache_path, "r", encoding="utf-8") as f:
        base_text = f.read()

    # Append vision-extracted content as additional context
    if vision_texts:
        additions = ["\n\n=== DIAGRAM AND FIGURE CONTENT (Vision Extracted) ===\n"]
        for page_idx, content in sorted(vision_texts.items()):
            additions.append(f"\n--- Page {page_idx+1} ---\n{content}")
        base_text += "\n".join(additions)

    print(f"  Merged text: {len(base_text):,} chars")
    return base_text

# ── Content generation ────────────────────────────────────────────────────────
FLASH_SYS = """MCCQE Part I tutor. Extract high-yield flashcards from Toronto Notes content.
Return ONLY JSON array. Each: {"specialty":"...","topic":"...","front":"...","back":"...","hy_flag":"Yes/No"}
front=term/concept/clinical trigger. back=explanation <=40 words.
Include drug doses, diagnostic criteria, anatomy relationships exactly as written.
If content describes a diagram or algorithm, create flashcards from the decision points.
Aim 8-15 per chunk."""

BULLET_SYS = """MCCQE Part I tutor. Extract condensed bullet notes from Toronto Notes content.
Return ONLY JSON array. Each: {"specialty":"...","topic":"...","note":"..."}
note=one standalone clinical fact <=25 words. Include exact numbers, drug names, doses, thresholds.
If content includes a table row, extract each row as a separate note.
If content describes an algorithm, extract each decision branch as a note.
Aim 10-20 per chunk."""

PARSE_MCQ_SYS = """Parse official MCCQE Part I practice questions verbatim.
Return ONLY JSON array. Each:
{"specialty":"...","question":"full stem","option_a":"...","option_b":"...","option_c":"...","option_d":"...","option_e":"...","correct_answer":"A-E","rationale":"..."}
Fill missing options with N/A."""

GEN_MCQ_SYS = """MCCQE Part I question writer. Generate original MCQs from Toronto Notes content.
Return ONLY JSON array. Each:
{"specialty":"...","question":"clinical vignette 3-5 sentences","option_a":"...","option_b":"...","option_c":"...","option_d":"...","option_e":"...","correct_answer":"A-E","rationale":"2 sentence explanation"}
3-5 questions per chunk. Canadian clinical context. Base facts strictly on provided content."""

DONTMISS_SYS = """MCCQE Part I tutor. Extract classic exam traps, buzzwords, and first-line treatments.
Return ONLY JSON array. Each: {"specialty":"...","item":"..."}
Format: Buzzword: X -> Y | Trap: A not B | First-line: condition -> drug/dose
Include anatomical relationships from diagrams if present.
5-10 per chunk."""


def generate_flashcards(client, chunks):
    rows = []
    for chunk in tqdm(chunks, desc="  Flashcards"):
        sp = detect_specialty(chunk)
        for item in safe_json(call_claude(client, FLASH_SYS, f"Specialty: {sp}\n\n{chunk}")):
            rows.append([item.get("specialty",sp), item.get("topic",""),
                         item.get("front",""), item.get("back",""), item.get("hy_flag","No")])
        time.sleep(0.8)
    return rows

def generate_bullets(client, chunks):
    rows = []
    for chunk in tqdm(chunks, desc="  Bullet Notes"):
        sp = detect_specialty(chunk)
        for item in safe_json(call_claude(client, BULLET_SYS, f"Specialty: {sp}\n\n{chunk}")):
            rows.append([item.get("specialty",sp), item.get("topic",""), item.get("note","")])
        time.sleep(0.8)
    return rows

def parse_official_mcqs(client, text):
    rows = []
    for chunk in tqdm(chunk_text(text, 12000), desc="  Official MCQs"):
        for item in safe_json(call_claude(client, PARSE_MCQ_SYS, chunk)):
            rows.append([item.get("specialty","General Medicine"), "A",
                         item.get("question",""),
                         item.get("option_a",""), item.get("option_b",""),
                         item.get("option_c",""), item.get("option_d",""),
                         item.get("option_e",""),
                         item.get("correct_answer",""), item.get("rationale","")])
        time.sleep(0.8)
    return rows

def generate_mcq_versions(client, chunks, versions):
    rows = []
    for version in versions:
        print(f"  Generating Version {version}...")
        sample = chunks[::max(1, len(chunks)//20)][:20]
        for chunk in tqdm(sample, desc=f"  Version {version}"):
            sp = detect_specialty(chunk)
            for item in safe_json(call_claude(client, GEN_MCQ_SYS,
                                              f"Specialty: {sp}\nVersion: {version}\n\n{chunk}")):
                rows.append([item.get("specialty",sp), version,
                             item.get("question",""),
                             item.get("option_a",""), item.get("option_b",""),
                             item.get("option_c",""), item.get("option_d",""),
                             item.get("option_e",""),
                             item.get("correct_answer",""), item.get("rationale","")])
            time.sleep(0.8)
    return rows

def generate_dont_miss(client, chunks):
    rows = []
    for chunk in tqdm(chunks, desc="  Don't Miss"):
        sp = detect_specialty(chunk)
        for item in safe_json(call_claude(client, DONTMISS_SYS, f"Specialty: {sp}\n\n{chunk}")):
            rows.append([item.get("specialty",sp), item.get("item","")])
        time.sleep(0.8)
    return rows

# ── Excel ─────────────────────────────────────────────────────────────────────
def write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    fc = HEADER_FILLS.get(name, "1F497D")
    fill = PatternFill("solid", start_color=fc, fgColor=fc)
    hf = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = hf
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    alt = PatternFill("solid", start_color="EBF3FB", fgColor="EBF3FB")
    bf = Font(name="Arial", size=10)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, str(val) if val else "")
            cell.font = bf
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r % 2 == 0:
                cell.fill = alt
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 60)
    ws.freeze_panes = "A2"

def build_excel(flash, bullets, mcq, dontmiss, path):
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Flashcards",   ["Specialty","Topic","Front","Back","HY_Flag"], flash)
    write_sheet(wb, "Bullet_Notes", ["Specialty","Topic","Note"], bullets)
    write_sheet(wb, "MCQ",          ["Specialty","Version","Question",
                                     "Option_A","Option_B","Option_C","Option_D","Option_E",
                                     "Correct_Answer","Rationale"], mcq)
    write_sheet(wb, "Dont_Miss",    ["Specialty","Item"], dontmiss)
    wb.save(path)

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--pdf1",       required=True,  help="Toronto Notes 2025 PDF")
    p.add_argument("--pdf2",       required=True,  help="MCC Practice Questions PDF")
    p.add_argument("--cache",      required=True,  help="OCR cache .txt file")
    p.add_argument("--key",        required=True,  help="Anthropic API key")
    p.add_argument("--output",     default="MCCQE_Study_Data.xlsx")
    p.add_argument("--images-dir", default="images")
    p.add_argument("--versions",   default="B,C,D")
    p.add_argument("--skip-scan",  action="store_true",
                   help="Skip diagram scan if diagram_pages.json exists")
    args = p.parse_args()

    client   = anthropic.Anthropic(api_key=args.key)
    versions = [v.strip().upper() for v in args.versions.split(",")]

    # ── Phase 1: find diagram pages ───────────────────────────────────────────
    diag_cache = "diagram_pages.json"
    if args.skip_scan and os.path.exists(diag_cache):
        print(f"\n[1/7] Loading diagram page list from {diag_cache}...")
        with open(diag_cache) as f:
            diagram_pages = [tuple(x) for x in json.load(f)]
    else:
        print("\n[1/7] Scanning for diagram pages...")
        diagram_pages = scan_for_diagram_pages(args.pdf1)
        with open(diag_cache, "w") as f:
            json.dump(diagram_pages, f)
        print(f"      Saved to {diag_cache}")

    # ── Phase 2: extract diagram images + Vision ──────────────────────────────
    img_index_cache = "image_index.json"
    if os.path.exists(img_index_cache):
        print(f"\n[2/7] Loading image index from {img_index_cache}...")
        with open(img_index_cache) as f:
            data = json.load(f)
        image_index = data["index"]
        vision_texts = {int(k): v for k, v in data.get("vision_texts", {}).items()}
    else:
        print(f"\n[2/7] Extracting diagram images + Vision processing...")
        image_index, vision_texts = extract_diagram_images(
            args.pdf1, diagram_pages, client, args.images_dir)
        with open(img_index_cache, "w") as f:
            json.dump({"index": image_index, "vision_texts": {str(k): v
                       for k, v in vision_texts.items()}}, f, indent=2)
        print(f"      Saved {len(image_index)} images to ./{args.images_dir}/")
        print(f"      Image index saved to {img_index_cache}")

    # ── Phase 3: build merged text ─────────────────────────────────────────────
    print("\n[3/7] Merging OCR cache + Vision content...")
    merged_text = build_merged_text(args.cache, vision_texts)
    tn_chunks   = chunk_text(merged_text)
    print(f"      {len(tn_chunks)} chunks")

    # ── Phase 4: MCC questions ─────────────────────────────────────────────────
    print("\n[4/7] Extracting MCC Practice Questions...")
    mcc_doc = fitz.open(args.pdf2)
    mcc_text = "\n\n".join(
        page.get_text() for page in mcc_doc if page.get_text().strip()
    )
    mcc_doc.close()
    print(f"      {len(mcc_text):,} chars")

    # ── Phase 5-8: generate content ────────────────────────────────────────────
    print("\n[5/7] Generating study content...")
    flash    = generate_flashcards(client, tn_chunks)
    bullets  = generate_bullets(client, tn_chunks)
    mcq      = parse_official_mcqs(client, mcc_text)
    mcq     += generate_mcq_versions(client, tn_chunks, versions)
    dontmiss = generate_dont_miss(client, tn_chunks)

    print(f"\n[6/7] Building Excel: {args.output}...")
    build_excel(flash, bullets, mcq, dontmiss, args.output)

    print(f"\n[7/7] Saving image index...")
    # Already saved above, just confirm
    print(f"      {len(image_index)} images indexed")

    # ── Summary ────────────────────────────────────────────────────────────────
    print("\n" + "="*60)
    print("SESSION 1 COMPLETE")
    print("="*60)
    print(f"Flashcards:    {len(flash):>5} rows")
    print(f"Bullet_Notes:  {len(bullets):>5} rows")
    print(f"MCQ:           {len(mcq):>5} rows  (A=official, {'/'.join(versions)}=AI)")
    print(f"Dont_Miss:     {len(dontmiss):>5} rows")
    print(f"Diagram images:{len(image_index):>5} files in ./{args.images_dir}/")
    print(f"\nFiles for Session 2:")
    print(f"  {args.output}")
    print(f"  {img_index_cache}")
    print(f"  ./{args.images_dir}/ (upload folder to Google Drive)")
    print("="*60)

if __name__ == "__main__":
    main()
