#!/usr/bin/env python3
"""
MCCQE Study App - Session 1 Pipeline (OCR Edition)
====================================================
Processes Toronto Notes 2025 (scanned PDF via OCR) +
MCC Practice Questions PDF (text-based).
Outputs structured Excel with 4 sheets for GAS web app (Session 2).

Usage:
    python mccqe_pipeline_v2.py \
        --pdf1 "Toronto Notes 2025.pdf" \
        --pdf2 "MCCQE-Part-I-practice-questions-and-answers-2025.pdf" \
        --key YOUR_ANTHROPIC_API_KEY

Output:
    MCCQE_Study_Data.xlsx
"""

import argparse
import json
import os
import re
import sys
import time

# ── dependency bootstrap ──────────────────────────────────────────────────────
def install(pkg):
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

for pkg, imp in [("anthropic","anthropic"),("pdfplumber","pdfplumber"),
                  ("openpyxl","openpyxl"),("tqdm","tqdm"),
                  ("pymupdf","fitz"),("pytesseract","pytesseract"),("pillow","PIL")]:
    try:
        __import__(imp)
    except ImportError:
        print(f"Installing {pkg}...")
        install(pkg)

import anthropic
import pdfplumber
import fitz
import pytesseract
from PIL import Image
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ── Tesseract path ────────────────────────────────────────────────────────────
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.exists(TESSERACT_PATH):
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

# ── constants ─────────────────────────────────────────────────────────────────
MODEL        = "claude-opus-4-5"
MAX_TOKENS   = 4096
CHUNK_CHARS  = 10000
RETRY_DELAY  = 15
MAX_RETRIES  = 3
OCR_DPI      = 200

SPECIALTIES = [
    "Cardiology","Respirology","Gastroenterology","Nephrology",
    "Hematology","Oncology","Endocrinology","Neurology",
    "Rheumatology","Infectious Disease","Dermatology",
    "Ophthalmology","Otolaryngology","Psychiatry","Obstetrics",
    "Gynecology","Pediatrics","Surgery","Orthopedics",
    "Emergency Medicine","Anesthesia","Radiology",
    "Family Medicine","Geriatrics","Pharmacology"
]

HEADER_FILLS = {
    "Flashcards":"1F497D","Bullet_Notes":"2E75B6",
    "MCQ":"375623","Dont_Miss":"7030A0"
}

# ── PDF extraction ────────────────────────────────────────────────────────────
def is_text_based(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        for i in range(min(5, len(pdf.pages))):
            t = pdf.pages[i].extract_text()
            if t and len(t.strip()) > 100:
                return True
    return False


def extract_text_pdf(pdf_path):
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for p in tqdm(pdf.pages, desc="  Reading pages"):
            t = p.extract_text()
            if t:
                pages.append(t.strip())
    return "\n\n".join(pages)


def extract_ocr_pdf(pdf_path):
    print("  Using OCR - this will take a while for large PDFs...")
    doc = fitz.open(pdf_path)
    total = len(doc)
    print(f"  Total pages: {total}")
    pages_text = []

    for i in tqdm(range(total), desc="  OCR pages"):
        page = doc[i]
        mat = fitz.Matrix(OCR_DPI / 72, OCR_DPI / 72)
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        text = pytesseract.image_to_string(img, lang="eng", config="--psm 6")
        if text.strip():
            pages_text.append(text.strip())

        if (i + 1) % 50 == 0:
            cache = pdf_path.replace(".pdf", f"_ocr_p{i+1}.txt")
            with open(cache, "w", encoding="utf-8") as f:
                f.write("\n\n".join(pages_text))
            print(f"\n  Partial save: {cache}")

    doc.close()
    full_text = "\n\n".join(pages_text)
    cache_path = pdf_path.replace(".pdf", "_ocr_cache.txt")
    with open(cache_path, "w", encoding="utf-8") as f:
        f.write(full_text)
    print(f"  OCR done. Cached: {cache_path}")
    return full_text


def extract_pdf(pdf_path):
    cache = pdf_path.replace(".pdf", "_ocr_cache.txt")
    if os.path.exists(cache):
        print(f"  Loading OCR cache: {cache}")
        with open(cache, "r", encoding="utf-8") as f:
            return f.read()
    if is_text_based(pdf_path):
        print("  Text-based PDF detected")
        return extract_text_pdf(pdf_path)
    print("  Scanned PDF detected, starting OCR")
    return extract_ocr_pdf(pdf_path)


def chunk_text(text, size=CHUNK_CHARS):
    return [text[i:i+size] for i in range(0, len(text), size)]


def detect_specialty(text):
    tl = text.lower()
    for s in SPECIALTIES:
        if s.lower() in tl:
            return s
    return "General Medicine"

# ── Claude API ────────────────────────────────────────────────────────────────
def call_claude(client, system, user):
    for attempt in range(MAX_RETRIES):
        try:
            r = client.messages.create(
                model=MODEL, max_tokens=MAX_TOKENS,
                system=system,
                messages=[{"role":"user","content":user}]
            )
            return r.content[0].text
        except anthropic.RateLimitError:
            wait = RETRY_DELAY * (attempt + 1)
            print(f"\n  Rate limit. Waiting {wait}s...")
            time.sleep(wait)
        except Exception as e:
            print(f"\n  API error: {e}")
            time.sleep(RETRY_DELAY)
    return "[]"


def safe_json(text):
    text = re.sub(r"```(?:json)?", "", text).strip().rstrip("`").strip()
    s, e = text.find("["), text.rfind("]")
    if s == -1 or e == -1:
        return []
    try:
        return json.loads(text[s:e+1])
    except:
        return []

# ── Content generators ────────────────────────────────────────────────────────
FLASH_SYS = """MCCQE Part I tutor. Extract high-yield flashcards from Toronto Notes.
Return ONLY JSON array. Each: {"specialty":"...","topic":"...","front":"...","back":"...","hy_flag":"Yes/No"}
front=term/concept, back=explanation <=40 words. Aim 8-15 per chunk."""

BULLET_SYS = """MCCQE Part I tutor. Extract condensed bullet notes from Toronto Notes.
Return ONLY JSON array. Each: {"specialty":"...","topic":"...","note":"..."}
note=one clinical fact <=25 words with numbers/drugs/criteria. Aim 10-20 per chunk."""

PARSE_MCQ_SYS = """Parse official MCCQE practice questions exactly as written.
Return ONLY JSON array. Each:
{"specialty":"...","question":"...","option_a":"...","option_b":"...","option_c":"...","option_d":"...","option_e":"...","correct_answer":"A/B/C/D/E","rationale":"..."}
Fill missing options with N/A."""

GEN_MCQ_SYS = """MCCQE Part I question writer. Write original MCQs from Toronto Notes content.
Return ONLY JSON array. Each:
{"specialty":"...","question":"clinical vignette 3-5 sentences","option_a":"...","option_b":"...","option_c":"...","option_d":"...","option_e":"...","correct_answer":"A/B/C/D/E","rationale":"2 sentences"}
Write 3-5 per chunk. Canadian clinical context."""

DONTMISS_SYS = """MCCQE Part I tutor. Extract classic exam traps and buzzwords from Toronto Notes.
Return ONLY JSON array. Each: {"specialty":"...","item":"..."}
Format: Buzzword/Trap/First-line: description. Aim 5-10 per chunk."""


def generate_flashcards(client, chunks):
    rows = []
    for chunk in tqdm(chunks, desc="  Flashcards"):
        sp = detect_specialty(chunk)
        for item in safe_json(call_claude(client, FLASH_SYS, f"Specialty: {sp}\n\n{chunk}")):
            rows.append([item.get("specialty",sp),item.get("topic",""),
                         item.get("front",""),item.get("back",""),item.get("hy_flag","No")])
        time.sleep(1)
    return rows


def generate_bullets(client, chunks):
    rows = []
    for chunk in tqdm(chunks, desc="  Bullet Notes"):
        sp = detect_specialty(chunk)
        for item in safe_json(call_claude(client, BULLET_SYS, f"Specialty: {sp}\n\n{chunk}")):
            rows.append([item.get("specialty",sp),item.get("topic",""),item.get("note","")])
        time.sleep(1)
    return rows


def parse_official_mcqs(client, text):
    rows = []
    for chunk in tqdm(chunk_text(text, 12000), desc="  Official MCQs"):
        for item in safe_json(call_claude(client, PARSE_MCQ_SYS, chunk)):
            rows.append([item.get("specialty","General Medicine"),"A",
                         item.get("question",""),
                         item.get("option_a",""),item.get("option_b",""),
                         item.get("option_c",""),item.get("option_d",""),
                         item.get("option_e",""),
                         item.get("correct_answer",""),item.get("rationale","")])
        time.sleep(1)
    return rows


def generate_mcq_versions(client, chunks, versions):
    rows = []
    for version in versions:
        print(f"  Generating Version {version}...")
        sample = chunks[::max(1, len(chunks)//20)][:20]
        for chunk in tqdm(sample, desc=f"  Version {version}"):
            sp = detect_specialty(chunk)
            for item in safe_json(call_claude(client, GEN_MCQ_SYS,
                                              f"Specialty: {sp}\nVersion: {version}\n\n{chunk}")):
                rows.append([item.get("specialty",sp),version,
                             item.get("question",""),
                             item.get("option_a",""),item.get("option_b",""),
                             item.get("option_c",""),item.get("option_d",""),
                             item.get("option_e",""),
                             item.get("correct_answer",""),item.get("rationale","")])
            time.sleep(1)
    return rows


def generate_dont_miss(client, chunks):
    rows = []
    for chunk in tqdm(chunks, desc="  Don't Miss"):
        sp = detect_specialty(chunk)
        for item in safe_json(call_claude(client, DONTMISS_SYS, f"Specialty: {sp}\n\n{chunk}")):
            rows.append([item.get("specialty",sp),item.get("item","")])
        time.sleep(1)
    return rows

# ── Excel builder ─────────────────────────────────────────────────────────────
def write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    fill = PatternFill("solid", start_color=HEADER_FILLS.get(name,"1F497D"),
                       fgColor=HEADER_FILLS.get(name,"1F497D"))
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = hfont
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    alt = PatternFill("solid", start_color="EBF3FB", fgColor="EBF3FB")
    bfont = Font(name="Arial", size=10)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, str(val) if val else "")
            cell.font = bfont
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r % 2 == 0:
                cell.fill = alt

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 60)
    ws.freeze_panes = "A2"


def build_excel(flash, bullets, mcq, dontmiss, path):
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Flashcards",  ["Specialty","Topic","Front","Back","HY_Flag"], flash)
    write_sheet(wb, "Bullet_Notes",["Specialty","Topic","Note"], bullets)
    write_sheet(wb, "MCQ",         ["Specialty","Version","Question",
                                    "Option_A","Option_B","Option_C","Option_D","Option_E",
                                    "Correct_Answer","Rationale"], mcq)
    write_sheet(wb, "Dont_Miss",   ["Specialty","Item"], dontmiss)
    wb.save(path)

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--pdf1",     required=True)
    p.add_argument("--pdf2",     required=True)
    p.add_argument("--key",      required=True)
    p.add_argument("--output",   default="MCCQE_Study_Data.xlsx")
    p.add_argument("--versions", default="B,C,D")
    args = p.parse_args()

    client   = anthropic.Anthropic(api_key=args.key)
    versions = [v.strip().upper() for v in args.versions.split(",")]

    print("\n[1/6] Extracting Toronto Notes...")
    tn_text   = extract_pdf(args.pdf1)
    tn_chunks = chunk_text(tn_text)
    print(f"      {len(tn_chunks)} chunks ({len(tn_text):,} chars)")

    print("\n[2/6] Extracting MCC Practice Questions...")
    mcc_text = extract_pdf(args.pdf2)
    print(f"      {len(mcc_text):,} chars")

    print("\n[3/6] Generating flashcards...")
    flash = generate_flashcards(client, tn_chunks)
    print(f"      {len(flash)} flashcards")

    print("\n[4/6] Generating bullet notes...")
    bullets = generate_bullets(client, tn_chunks)
    print(f"      {len(bullets)} notes")

    print("\n[5/6] Processing MCQs...")
    mcq  = parse_official_mcqs(client, mcc_text)
    mcq += generate_mcq_versions(client, tn_chunks, versions)
    print(f"      {len(mcq)} MCQ rows")

    print("\n[6/6] Generating Don't Miss items...")
    dontmiss = generate_dont_miss(client, tn_chunks)
    print(f"      {len(dontmiss)} items")

    print(f"\nBuilding {args.output}...")
    build_excel(flash, bullets, mcq, dontmiss, args.output)

    print("\n" + "="*60)
    print("SESSION 1 COMPLETE - SCHEMA FOR SESSION 2")
    print("="*60)
    print(f"Flashcards:   {len(flash):>5} rows | Specialty,Topic,Front,Back,HY_Flag")
    print(f"Bullet_Notes: {len(bullets):>5} rows | Specialty,Topic,Note")
    print(f"MCQ:          {len(mcq):>5} rows | Specialty,Version,Question,Options,Answer,Rationale")
    print(f"Dont_Miss:    {len(dontmiss):>5} rows | Specialty,Item")
    print(f"\nMCQ Versions: A=official MCC, {'/'.join(versions)}=AI generated")
    print(f"Output: {args.output}")
    print("="*60)


if __name__ == "__main__":
    main()
